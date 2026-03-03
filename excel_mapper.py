import pandas as pd
import openpyxl
import io
import zipfile

def extract_value_from_excel(file_bytes, sheet_name, cell_coordinate):
    """
    Extract a single value from an Excel file bytes stream, ignoring formulas.
    Returns None if sheet or cell missing.
    """
    try:
        # data_only=True ensures we read the computed value, not the formula string.
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        if sheet_name not in wb.sheetnames:
            return None
        sheet = wb[sheet_name]
        return sheet[cell_coordinate].value
    except Exception as e:
        print(f"Error reading source file: {e}")
        return None

def write_values_to_template(template_bytes, mappings):
    """
    Opens the template file, writes the collected values to their respective sheets/cells,
    and returns the new file bytes.
    Mappings format: list of dicts:
    [{'Target Sheet': '...', 'Target Cell': '...', 'Value': ...}]
    """
    wb = openpyxl.load_workbook(io.BytesIO(template_bytes))
    for mapping in mappings:
        sheet_name = mapping.get('Target Report Sheet')
        cell_coord = mapping.get('Target Report Cell')
        value = mapping.get('Value')
        
        if sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            sheet[cell_coord].value = value
            
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

def process_mapping_execution(mapping_df, source_files_dict, template_files_dict):
    """
    Executes the report generation based on the config.
    Returns: Dict of {generated_file_name: bytes} or raises ValueError
    """
    # Group the mapping by Target Report Name, so we only process each template once per generation round.
    # A generated file might just be the exact same Target Report Name, but we will output it directly.
    generated_reports = {}
    
    # We will aggregate all the needed actions per Target Report
    target_tasks = {}
    
    for index, row in mapping_df.iterrows():
        source_file = str(row.get('Source File Name', '')).strip()
        source_sheet = str(row.get('Source Sheet', '')).strip()
        source_cell = str(row.get('Source Cell', '')).strip()
        
        target_file = str(row.get('Target Report File Name', '')).strip()
        target_sheet = str(row.get('Target Report File Sheet', '')).strip()
        target_cell = str(row.get('Target Report File Cell', '')).strip()
        
        if not all([source_file, source_sheet, source_cell, target_file, target_sheet, target_cell]):
            continue # skip empty or invalid rows
            
        # 1. Fetch source value
        if source_file not in source_files_dict:
            raise ValueError(f"Source file '{source_file}' was not uploaded, but required in mapping.")
            
        value = extract_value_from_excel(source_files_dict[source_file], source_sheet, source_cell)
        
        # 2. Add to target tasks
        if target_file not in target_tasks:
            target_tasks[target_file] = []
            
        target_tasks[target_file].append({
            'Target Report Sheet': target_sheet,
            'Target Report Cell': target_cell,
            'Value': value
        })
        
    # 3. Render all final templates
    for target_file, tasks in target_tasks.items():
        if target_file not in template_files_dict:
            raise ValueError(f"Target template '{target_file}' was not uploaded, but required in mapping.")
            
        final_bytes = write_values_to_template(template_files_dict[target_file], tasks)
        generated_reports[f"GENERATED_{target_file}"] = final_bytes
        
    return generated_reports

def create_zip_archive(files_dict):
    """
    Combine multiple file bytes into a single ZIP file bytes buffer.
    """
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
        for file_name, file_bytes in files_dict.items():
            zip_file.writestr(file_name, file_bytes)
    zip_buffer.seek(0)
    return zip_buffer.getvalue()

def generate_mock_mapping_file():
    """
    Helper function to provide the user with a blank mapping template.
    """
    df = pd.DataFrame(columns=[
        'Source File Name', 
        'Source Sheet', 
        'Source Cell',
        'Target Report File Name',
        'Target Report File Sheet',
        'Target Report File Cell'
    ])
    # add one row example
    df.loc[0] = ['data1.xlsx', 'Sheet1', 'A2', 'report_template.xlsx', 'Summary', 'B4']
    
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Mapping')
    output.seek(0)
    return output.getvalue()
